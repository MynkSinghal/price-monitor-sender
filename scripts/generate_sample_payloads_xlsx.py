"""Generate a side-by-side Excel preview of what UKPROD and USPROD will POST.

Three sheets: README (cover), UKPROD, USPROD. Each site sheet shows the CSV
grid that the sender will produce in a happy-path production scenario where
every job local to that site has completed (cross-site / audit-only rows
are left blank, exactly as on the wire).

Usage:
    .venv/bin/python -m scripts.generate_sample_payloads_xlsx
    -> writes price_monitor_sender/sample_payloads.xlsx
"""

from __future__ import annotations

import importlib
import os
from datetime import datetime, timedelta
from pathlib import Path

os.environ.setdefault("RECEIVER_URL", "http://example/prices")

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

PROJECT_ROOT = Path(__file__).resolve().parent.parent
OUT_PATH = PROJECT_ROOT / "sample_payloads.xlsx"


# =====================================================================
#              PALETTE — DARK MODE / ION TRADING GROUP
# =====================================================================
# ION's identity: near-black surfaces with electric cyan as the single
# brand accent. Everything else is monochrome with a subtle cool tint.
BG_PAGE       = "0B0E13"   # page / sheet background (true black-ish)
BG_TITLE      = "000000"   # title banner — pure black
BG_HEADER     = "12171F"   # column-header row (slightly lifted)
BG_ROW        = "161B22"   # default body row (lifted "card" feel)
BG_ROW_ALT    = "1A2029"   # zebra alternate
BG_KPI        = "0F1218"   # KPI strip background
BG_LEGEND     = "12171F"   # legend block

ION_CYAN      = "00D4FF"   # signature accent — used SPARINGLY on dark
ION_CYAN_DIM  = "0098B5"   # darker cyan for borders / muted accents
WHITE         = "FFFFFF"
TEXT_PRIMARY  = "E8EDF3"   # 92% white — body text
TEXT_MUTED    = "8B95A3"   # secondary / placeholders / italic em-dash
TEXT_DIM      = "5A6470"   # tertiary

# Category fills — muted on dark, bright "glow" pill colour for accent.
# (body fill is dark, only the row_type pill + left edge bar carry the
#  vibrant colour — keeps the sheet calm + scannable)
FILL_GREEN_BG    = "16271F"   # filled body — deep teal-tinted dark
FILL_GREEN_BAR   = "00E5A8"   # filled accent — mint glow
FILL_PEACH_BG    = "2A1A12"   # cross-site body — warm dark
FILL_PEACH_BAR   = "FF8A4C"   # cross-site accent — amber glow
FILL_YELLOW_BG   = "2A2410"   # audit body — gold-tinted dark
FILL_YELLOW_BAR  = "FFD93D"   # audit accent — gold glow
FILL_BLUE_BG     = "10202E"   # OR-mode body — deep navy
FILL_BLUE_BAR    = "5EA8FF"   # OR-mode accent — sky blue glow

HEADER_FILL = PatternFill("solid", fgColor=BG_HEADER)
TITLE_FILL = PatternFill("solid", fgColor=BG_TITLE)
ACCENT_FILL = PatternFill("solid", fgColor=ION_CYAN)
PAGE_FILL = PatternFill("solid", fgColor=BG_PAGE)


def _load_for_site(site: str):
    os.environ["SENDER_SITE"] = site
    import src.config_loader as cl
    importlib.reload(cl)
    return cl.load_config()


def _seeded_timestamp(idx: int) -> str:
    base = datetime(2026, 4, 23, 6, 0, 0)
    return (base + timedelta(minutes=idx * 5, seconds=(idx * 7) % 60)).strftime(
        "%d/%m/%Y %H:%M:%S"
    )


# Cell styling helpers -------------------------------------------------

# Hairline borders in a slightly-lighter-than-bg tone so rows stay
# distinct without being noisy on dark.
THIN = Side(border_style="thin", color="232A33")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ROW_HEIGHT = 22
HEADER_ROW_HEIGHT = 36
TITLE_ROW_HEIGHT = 56


def _paint_surrounding_dark(ws, *, last_data_row: int, last_data_col: int,
                            extra_cols: int = 14, extra_rows: int = 8) -> None:
    """Extend the dark-mode background to the right and below the data area
    so the sheet looks fully dark at typical zoom levels (Excel has no native
    'sheet background fill' API; we paint cells instead)."""
    for col in range(last_data_col + 1, last_data_col + 1 + extra_cols):
        for row in range(1, last_data_row + extra_rows + 1):
            ws.cell(row=row, column=col).fill = PAGE_FILL
    for row in range(last_data_row + 1, last_data_row + 1 + extra_rows):
        for col in range(1, last_data_col + 1 + extra_cols):
            ws.cell(row=row, column=col).fill = PAGE_FILL


def _categorise(pg, cross_site_jobs):
    """Return (row_type_label, body_fill_hex, bar_fill_hex, has_timestamp)."""
    if not pg.jobs:
        return ("AUDIT-ONLY", FILL_YELLOW_BG, FILL_YELLOW_BAR, False)
    jobset = set(pg.jobs)
    if jobset.issubset(cross_site_jobs):
        return ("CROSS-SITE", FILL_PEACH_BG, FILL_PEACH_BAR, False)
    if pg.match_mode == "any":
        return ("OR-MODE", FILL_BLUE_BG, FILL_BLUE_BAR, True)
    if pg.is_composite:
        return ("COMPOSITE", FILL_GREEN_BG, FILL_GREEN_BAR, True)
    return ("SINGLE-JOB", FILL_GREEN_BG, FILL_GREEN_BAR, True)


def _populate_sheet(ws, cfg, site_label: str) -> None:
    cross_site = cfg.cross_site_jobs

    # ---- counts (used in title banner) ----
    audit = sum(1 for pg in cfg.active_price_groups if not pg.jobs)
    cross_empty = sum(
        1 for pg in cfg.active_price_groups
        if pg.jobs and set(pg.jobs).issubset(cross_site)
    )
    filled = len(cfg.active_price_groups) - audit - cross_empty
    total = len(cfg.active_price_groups)

    headers = [
        "#",
        "price_group_name",
        "timestamp",
        "raw_csv_line",
        "row_type",
        "match_mode",
        "jobs",
        "notes",
    ]
    n_cols = len(headers)
    last_col = get_column_letter(n_cols)

    # ====== ROW 1 — TITLE BANNER (pure black with cyan ION mark) ======
    ws.merge_cells(f"A1:{last_col}1")
    cell = ws["A1"]
    cell.value = (
        f"  ION  ◆  PRICE MONITOR     {site_label}     ›     CSV PAYLOAD PREVIEW  "
    )
    cell.font = Font(name="Calibri", size=24, bold=True, color=WHITE)
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = TITLE_ROW_HEIGHT

    # ====== ROW 2 — single cyan accent stripe (signature ION mark) ======
    ws.row_dimensions[2].height = 4
    for col in range(1, n_cols + 1):
        ws.cell(row=2, column=col).fill = ACCENT_FILL

    # ====== ROW 3 — KPI strip (dark cards with coloured glyph) ======
    ws.row_dimensions[3].height = 36

    def _kpi(col_start: int, col_end: int, label: str, value: str, bar_color: str) -> None:
        ws.merge_cells(
            start_row=3, start_column=col_start, end_row=3, end_column=col_end
        )
        c = ws.cell(row=3, column=col_start)
        c.value = f"  {value}    {label}"
        # value is bold-bright, label is muted — single line, single cell.
        c.font = Font(name="Calibri", size=12, bold=True, color=bar_color)
        c.fill = PatternFill("solid", fgColor=BG_KPI)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = Border(
            left=Side(border_style="thick", color=bar_color),
            top=Side(border_style="thin", color="232A33"),
            bottom=Side(border_style="thin", color="232A33"),
            right=Side(border_style="thin", color="232A33"),
        )

    _kpi(1, 2, "TOTAL ROWS", str(total), ION_CYAN)
    _kpi(3, 4, "FILLED HERE", str(filled), FILL_GREEN_BAR)
    _kpi(5, 5, "CROSS-SITE", str(cross_empty), FILL_PEACH_BAR)
    _kpi(6, 6, "AUDIT-ONLY", str(audit), FILL_YELLOW_BAR)
    _kpi(7, 8, "OR-MODE",
         str(sum(1 for pg in cfg.active_price_groups if pg.match_mode == "any")),
         FILL_BLUE_BAR)

    # ====== ROW 4 — second cyan accent stripe ======
    ws.row_dimensions[4].height = 4
    for col in range(1, n_cols + 1):
        ws.cell(row=4, column=col).fill = ACCENT_FILL

    # ====== ROW 5 — column headers ======
    header_row = 5
    ws.row_dimensions[header_row].height = HEADER_ROW_HEIGHT
    for col_idx, name in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=name.upper())
        c.font = Font(name="Calibri", size=10, bold=True, color=ION_CYAN)
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(
            top=Side(border_style="thin", color="232A33"),
            bottom=Side(border_style="medium", color=ION_CYAN_DIM),
            left=THIN,
            right=THIN,
        )

    # ====== DATA ROWS ======
    body_alignment_center = Alignment(horizontal="center", vertical="center")
    body_alignment_left = Alignment(horizontal="left", vertical="center", indent=1)
    raw_alignment = Alignment(horizontal="left", vertical="center", indent=1)
    notes_alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=False)

    filled_idx = 0
    for i, pg in enumerate(cfg.active_price_groups, start=1):
        row_idx = header_row + i
        ws.row_dimensions[row_idx].height = ROW_HEIGHT
        row_type, body_fill_hex, bar_hex, has_ts = _categorise(pg, cross_site)

        timestamp = ""
        if has_ts:
            filled_idx += 1
            timestamp = _seeded_timestamp(filled_idx)

        raw_line = f"{pg.price_group_name}|{timestamp}"

        # Per-row text colours — primary white-ish on dark, accent on the
        # row_type pill, muted cyan on the price-group-name (brand touch).
        cells = [
            (1, i,                       body_alignment_center, "Calibri",  10, False, TEXT_DIM),
            (2, pg.price_group_name,     body_alignment_left,   "Calibri",  11, True,  ION_CYAN),
            (3, timestamp,               body_alignment_center, "Consolas", 11, True,  TEXT_PRIMARY),
            (4, raw_line,                raw_alignment,         "Consolas", 10, False, TEXT_PRIMARY),
            (5, row_type,                body_alignment_center, "Calibri",  10, True,  WHITE),  # pill
            (6, pg.match_mode,           body_alignment_center, "Calibri",  10, False, TEXT_MUTED),
            (7, ", ".join(pg.jobs),      body_alignment_left,   "Consolas",  9, False, TEXT_MUTED),
            (8, pg.notes or "",          notes_alignment,       "Calibri",   9, False, TEXT_MUTED),
        ]

        # Body fill: zebra-striped between two near-black tones, with a
        # subtle category tint on the timestamp column for filled rows.
        is_alt = (i % 2 == 0)
        base_bg = BG_ROW_ALT if is_alt else BG_ROW
        body_fill = PatternFill("solid", fgColor=base_bg)
        # Category-tinted body for the leftmost columns to keep the
        # category visible at a glance — but the rest of the row stays
        # the calm dark base.
        category_fill = PatternFill("solid", fgColor=body_fill_hex)

        for col, value, align, font_name, size, bold, color in cells:
            c = ws.cell(row=row_idx, column=col, value=value)
            # Columns 1..3 carry the category tint; 4..8 stay dark base.
            if col in (1, 3):
                c.fill = category_fill
            else:
                c.fill = body_fill
            c.alignment = align
            c.border = BORDER
            italic = (col == 8)
            c.font = Font(
                name=font_name, size=size, bold=bold, italic=italic, color=color
            )

        # row_type column gets a vivid solid pill (overrides body fill)
        pill = ws.cell(row=row_idx, column=5)
        pill.fill = PatternFill("solid", fgColor=bar_hex)
        # Use black text on bright pills for readability (cyan/yellow/mint)
        # and white on the warmer ones (peach/blue).
        pill_text = "0B0E13" if bar_hex in (FILL_GREEN_BAR, FILL_YELLOW_BAR, ION_CYAN) else WHITE
        pill.font = Font(name="Calibri", size=10, bold=True, color=pill_text)
        pill.border = Border(
            left=Side(border_style="thin", color=bar_hex),
            right=Side(border_style="thin", color=bar_hex),
            top=Side(border_style="thin", color=bar_hex),
            bottom=Side(border_style="thin", color=bar_hex),
        )

        # Left-edge accent: thick coloured left border on column 1
        edge = ws.cell(row=row_idx, column=1)
        edge.border = Border(
            left=Side(border_style="thick", color=bar_hex),
            right=THIN,
            top=THIN,
            bottom=THIN,
        )

        # Blank-timestamp gets a dim italic placeholder.
        if not timestamp:
            ts_cell = ws.cell(row=row_idx, column=3, value="—")
            ts_cell.font = Font(
                name="Calibri", size=11, italic=True, color=TEXT_DIM
            )
            ts_cell.alignment = body_alignment_center
            ts_cell.fill = category_fill
            ts_cell.border = BORDER

    # ====== column widths ======
    widths = [6, 50, 22, 60, 16, 12, 38, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ====== freeze + autofilter ======
    ws.freeze_panes = ws.cell(row=header_row + 1, column=2)
    ws.auto_filter.ref = (
        f"A{header_row}:{last_col}{header_row + len(cfg.active_price_groups)}"
    )

    # ====== sheet view (dark-mode niceties) ======
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 100

    # ====== legend block at the bottom ======
    legend_start = header_row + len(cfg.active_price_groups) + 2
    legend = [
        ("LEGEND", None, ION_CYAN, True),
        ("FILLED  ›  this sender has the data, timestamp on the wire", None, FILL_GREEN_BAR, False),
        ("CROSS-SITE  ›  the OTHER sender fills this row, blank here", None, FILL_PEACH_BAR, False),
        ("AUDIT-ONLY  ›  both senders blank, receiver fills manually", None, FILL_YELLOW_BAR, False),
        ("OR-MODE  ›  match_mode=any (e.g. JSE1/JSE), flags on first arrival", None, FILL_BLUE_BAR, False),
    ]
    for offset, (text, _, color, is_header) in enumerate(legend):
        r = legend_start + offset
        ws.row_dimensions[r].height = 24
        c1 = ws.cell(row=r, column=1)
        c1.fill = PatternFill("solid", fgColor=color)
        c1.border = Border(
            left=Side(border_style="thick", color=color),
            top=THIN, bottom=THIN, right=THIN,
        )
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=n_cols)
        c2 = ws.cell(row=r, column=2, value=text)
        if is_header:
            c2.font = Font(name="Calibri", size=12, bold=True, color=ION_CYAN)
        else:
            c2.font = Font(name="Calibri", size=10, bold=False, color=TEXT_PRIMARY)
        c2.fill = PatternFill("solid", fgColor=BG_LEGEND)
        c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c2.border = Border(top=THIN, bottom=THIN, right=THIN)

    # Final flourish: paint everything beyond the data dark
    last_row = legend_start + len(legend) - 1
    _paint_surrounding_dark(ws, last_data_row=last_row, last_data_col=n_cols)


def _add_readme_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("README", 0)
    ws.sheet_view.showGridLines = False

    # column widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 110
    ws.column_dimensions["C"].width = 4

    # Title block — pure black with cyan ION mark
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 64
    ws.row_dimensions[3].height = 6
    ws.row_dimensions[4].height = 22

    ws.merge_cells("A1:C1")
    ws["A1"].fill = PAGE_FILL

    ws.merge_cells("A2:C2")
    title = ws["A2"]
    title.value = "  ION  ◆  PRICE MONITOR  —  CSV PAYLOAD PREVIEW"
    title.font = Font(name="Calibri", size=30, bold=True, color=WHITE)
    title.fill = TITLE_FILL
    title.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A3:C3")
    ws["A3"].fill = ACCENT_FILL

    ws.merge_cells("A4:C4")
    sub = ws["A4"]
    sub.value = "  what each sender will POST in production"
    sub.font = Font(name="Calibri", size=12, italic=True, color=ION_CYAN)
    sub.fill = PAGE_FILL
    sub.alignment = Alignment(horizontal="left", vertical="center")

    sections = [
        ("WHAT THIS WORKBOOK SHOWS",
         "Two sheets — UKPROD and USPROD — each with 187 rows of the live CSV the sender will POST every minute.",
         "Both senders POST to the same automation endpoint. Both contain the same row set, in the same order. The only difference is which side fills in the timestamp."),
        ("HOW TO READ A SHEET",
         "Each row is one line in the POST body. The 'RAW_CSV_LINE' column is exactly what goes on the wire (e.g. PATH|23/04/2026 17:05:33 or RCFT|).",
         "Header strip up top has TOTAL ROWS, FILLED HERE, CROSS-SITE, AUDIT-ONLY, OR-MODE counters. Frozen panes + autofilter keep navigation fast."),
        ("ROW TYPES",
         "FILLED  ›  this sender has the success.txt; timestamp is filled.",
         "CROSS-SITE  ›  job lives on the OTHER sender; blank here. Receiver merges from the other site's POST.\nAUDIT-ONLY  ›  no RANTask job at all (KSE clients, manual-fill prices, ISTM, PSTM/SSTM/POMT/SOMT, RCFT, PCFF, PDCE). Both senders blank; receiver fills manually.\nOR-MODE  ›  match_mode=any. Currently only JSE1/JSE — flags as soon as either JSE or JSE1 arrives."),
        ("ABOUT THE TIMESTAMPS IN THIS PREVIEW",
         "Synthetic — spread across 23/04/2026 starting 06:00 IST so the sheet is readable.",
         "In production each row carries the real st_ctime of the corresponding success.txt file — the sender does no transformation."),
        ("REGENERATING THIS FILE",
         "Anytime you edit price_groups.json, re-run:",
         ".venv/bin/python -m scripts.generate_sample_payloads_xlsx"),
    ]

    row = 6
    for header, line1, line2 in sections:
        # Spacer row above header — tinted dark
        ws.row_dimensions[row - 1].height = 8
        for col_letter in ("A", "B", "C"):
            ws[f"{col_letter}{row - 1}"].fill = PAGE_FILL

        ws.row_dimensions[row].height = 30
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        h = ws.cell(row=row, column=1, value=f"  {header}")
        h.font = Font(name="Calibri", size=14, bold=True, color=ION_CYAN)
        h.fill = PatternFill("solid", fgColor=BG_HEADER)
        h.alignment = Alignment(horizontal="left", vertical="center")
        h.border = Border(left=Side(border_style="thick", color=ION_CYAN))
        row += 1

        for line in (line1, line2):
            for sub_text in line.split("\n"):
                ws.row_dimensions[row].height = 22
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
                c = ws.cell(row=row, column=1, value=f"   {sub_text}")
                c.font = Font(name="Calibri", size=11, color=TEXT_PRIMARY)
                c.fill = PatternFill("solid", fgColor=BG_ROW)
                c.alignment = Alignment(horizontal="left", vertical="center")
                row += 1
        row += 1  # gap

    # Pad some empty dark rows below for the dark-mode "page" feel
    for r in range(row, row + 12):
        ws.row_dimensions[r].height = 18
        for col in (1, 2, 3):
            ws.cell(row=r, column=col).fill = PAGE_FILL

    _paint_surrounding_dark(ws, last_data_row=row + 12, last_data_col=3, extra_cols=20)

    ws.sheet_properties.tabColor = ION_CYAN


def _set_tab_colors(wb: Workbook) -> None:
    if "UKPROD" in wb.sheetnames:
        wb["UKPROD"].sheet_properties.tabColor = ION_CYAN
    if "USPROD" in wb.sheetnames:
        wb["USPROD"].sheet_properties.tabColor = "FF6B35"  # warm orange counterpart


def main() -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for site in ("UKPROD", "USPROD"):
        cfg = _load_for_site(site)
        ws = wb.create_sheet(site)
        ws.sheet_view.showGridLines = False
        _populate_sheet(ws, cfg, site)

    _add_readme_sheet(wb)
    _set_tab_colors(wb)

    wb.active = wb.sheetnames.index("README")
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
    print(f"Sheets: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
